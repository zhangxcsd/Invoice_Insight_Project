"""ODS 层处理器 - 原始数据导入模块

=== ODS 层职责（Operational Data Store）===
ODS 层是数据仓库的"贴源层"，负责将原始 Excel 文件的数据原样导入数据库，
不做任何业务逻辑处理，仅做必要的数据类型转换和审计字段添加。

**核心目标：**
- 数据可追溯：保留原始数据，记录来源文件、导入时间等审计信息
- 快速导入：支持并行处理多个文件，优化大批量导入性能
- 容错处理：单个文件失败不影响其他文件的处理
- 灵活扩展：支持新增工作表类型和数据源

=== 处理流程 ===

**1. 文件扫描（Scan Phase）**
   ```
   输入：Excel 文件目录
   输出：文件清单 manifest (List[Dict])
   
   manifest 结构：
   {
       'filepath': '绝对路径',
       'filename': '文件名',
       'sheetnames': ['Sheet1', 'Sheet2'],
       'columns': {'Sheet1': ['列A', '列B'], ...},
       'source_category': 'purchase'/'sales',  # 进项/销项
       'year': 2024
   }
   ```
   
**2. 工作表分类（Sheet Classification）**
   根据工作表名称和列名识别表类型：
   - **detail 表**：包含"发票代码"、"发票号码"、"开票日期"等列
   - **header 表**：包含"销方名称"、"购方名称"等汇总信息
   - **summary 表**：汇总统计表，通常包含"合计"字样
   - **special 表**：特殊业务表，如"不动产"、"机动车"等
   
**3. 数据导入（Import Phase）**
   ```
   流程：Excel → ODS_*_TEMP_TRANSIT → ODS_* 正式表
   
   临时表的作用：
   - 批量数据验证（检查必填字段、数据类型等）
   - 批量插入优化（executemany）
   - 原子性保证（要么全成功，要么全失败）
   ```

**4. 并行处理（Parallel Processing）**
   - 使用 multiprocessing 模块实现多进程并行
   - worker_count 可配置（默认 CPU 核心数 - 1）
   - 每个 worker 独立处理一个文件，失败不影响其他
   - 自动内存监控，内存不足时降级为串行处理

=== 数据规范化 ===
虽然 ODS 层保持原始数据，但会做最小化的规范化处理：
- **日期格式**：统一转换为 YYYY-MM-DD
- **金额格式**：去除千分位逗号，保留两位小数
- **字符串清理**：去除前后空白，统一空值为 NULL
- **类型转换**：根据配置进行必要的类型转换（str → float, str → date）

=== 审计字段 ===
每条记录会自动添加以下审计字段：
- `source_file`: 来源文件名
- `source_filepath`: 来源文件绝对路径
- `source_sheet`: 来源工作表名
- `import_timestamp`: 导入时间（格式：YYYY-MM-DD HH:MM:SS）
- `year`: 所属年份（从文件名或配置中提取）
- `source_category`: 数据类别（purchase 进项 / sales 销项）

=== 错误处理 ===
- **文件级错误**：文件无法打开、格式错误等，记录到 error_logs.csv
- **工作表级错误**：工作表不符合预期格式，跳过该表并记录
- **行级错误**：单行数据解析失败，跳过该行并记录到 conversion_failures
- **数据库错误**：插入失败自动回滚到导入前状态

=== 性能优化策略 ===
1. **流式读取**：使用 openpyxl 只读模式，避免大文件内存溢出
2. **批量插入**：使用 executemany 减少数据库往返次数
3. **并行处理**：多文件同时导入，充分利用多核 CPU
4. **内存监控**：实时监控内存使用，超过阈值自动降级
5. **临时文件**：避免在内存中构建完整数据结构

=== 典型使用场景 ===
```python
# 场景 1：导入单个目录的所有文件
from vat_audit_pipeline.processors.ods_processor import process_ods_layer

manifest = scan_files(source_dir)
stats = process_ods_layer(db_conn, manifest, config)
print(f"成功导入 {stats['success']} 个文件，失败 {stats['failed']} 个")

# 场景 2：并行导入大批量文件
config['worker_count'] = 4  # 使用 4 个进程
stats = process_ods_layer(db_conn, large_manifest, config)

# 场景 3：仅扫描不导入（用于预检查）
manifest = scan_files(source_dir)
print(f"发现 {len(manifest)} 个文件，共 {sum(len(m['sheetnames']) for m in manifest)} 个工作表")
```

=== 维护建议 ===
1. **新增表类型**：在 classify_sheet() 函数中添加识别逻辑
2. **调整并行度**：根据硬件配置调整 worker_count（CPU 密集型任务建议 CPU 核心数 - 1）
3. **排查错误**：查看 error_logs.csv 和 vat_audit.log
4. **性能调优**：调整 batch_size（默认 1000）和 memory_threshold（默认 80%）
"""

from __future__ import annotations

import multiprocessing
import os
import shutil
import sqlite3
import uuid
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import pandas as pd
from openpyxl import load_workbook

from vat_audit_pipeline.core import models
from vat_audit_pipeline.core.models import RuntimeContext
from vat_audit_pipeline.utils.file_handlers import (
    add_audit_columns,
    add_dedup_capture_time,
    add_invoice_year_column,
    cleanup_old_temp_files,
    cleanup_temp_files,
    ensure_worker_temp_dir,
    filter_dataframe_columns,
    format_timestamp_for_filename,
    generate_manifest_filename,
    register_cleanup,
    save_dataframe_to_csv,
)
from vat_audit_pipeline.utils.parallel import calculate_optimal_workers, measure_disk_busy_percent
from vat_audit_pipeline.utils.validators import validate_input_file
from vat_audit_pipeline.utils.encoding import read_csv_with_encoding_detection
from vat_audit_pipeline.utils.logging import MemoryMonitor, PerformanceTimer, _debug_var, _progress
from vat_audit_pipeline.utils.normalization import cast_and_record as _cast_and_record
from vat_audit_pipeline.utils.sheet_processing import (
    PipelineSettings,
    SheetProcessingContext,
    SheetTypeMapping,
    get_sheet_handler,
    normalize_sheet_dataframe,
    write_to_csv_or_queue,
)


class ODSProcessor:
    """Lightweight OO wrapper over the existing ODS functions.

    This class is intentionally a thin adapter to support gradual refactors.
    The canonical implementation remains the module-level `process_ods`.
    """

    def __init__(self, runtime: RuntimeContext, logger, config: Optional[Any], conn: sqlite3.Connection):
        self.runtime = runtime
        self.logger = logger
        self.config = config
        self.conn = conn

    def process(
        self,
        excel_files: List[str],
        files_meta: Dict[str, Any],
        detail_columns: List[str],
        header_columns: List[str],
        summary_columns: List[str],
        special_columns: Dict[str, List[str]],
        process_time: str,
    ) -> Dict[str, Any]:
        return process_ods(
            self.runtime,
            self.logger,
            excel_files,
            files_meta,
            detail_columns,
            header_columns,
            summary_columns,
            special_columns,
            process_time,
            self.config,
            self.conn,
        )


def read_excel_with_engine(
    file_path: str,
    sheet_name: Optional[str] | Optional[int] | Optional[List[str]] = None,
    **kwargs: Any,
) -> pd.DataFrame | Dict[str, pd.DataFrame]:
    engine = "xlrd" if str(file_path).lower().endswith(".xls") else None
    return pd.read_excel(file_path, sheet_name=sheet_name, engine=engine, **kwargs)


def should_use_streaming_for_file(file_path: str, config: Optional[Any] = None) -> bool:
    try:
        import psutil

        file_size_mb = os.path.getsize(file_path) / (1024 * 1024)
        large_file_threshold_mb = 100
        stream_threshold_percent = 75

        if config and hasattr(config, "get"):
            large_file_threshold_mb = config.get(
                "performance",
                "memory_monitoring",
                "large_file_streaming_mb",
                default=large_file_threshold_mb,
            )
            stream_threshold_percent = config.get(
                "performance",
                "memory_monitoring",
                "stream_switch_threshold_percent",
                default=stream_threshold_percent,
            )

        if file_size_mb > large_file_threshold_mb:
            return True

        system_memory_percent = psutil.virtual_memory().percent
        if system_memory_percent >= stream_threshold_percent:
            return True

        available_mb = psutil.virtual_memory().available / (1024 * 1024)
        if file_size_mb > available_mb * 0.4:
            return True

        return False
    except Exception:
        return False


def cast_and_record(df: pd.DataFrame, fname: str, sheet: str, cast_stats, cast_failures, tax_text_to_zero: bool) -> pd.DataFrame:
    return _cast_and_record(df, fname, sheet, cast_stats, cast_failures, tax_text_to_zero=tax_text_to_zero)


def export_debug_samples(conn: sqlite3.Connection, tables: List[str], output_dir: str, limit: int = 100) -> None:
    if not tables:
        return
    debug_dir = os.path.join(output_dir, "debug_samples")
    os.makedirs(debug_dir, exist_ok=True)
    for table in tables:
        try:
            df_sample = pd.read_sql_query(f"SELECT * FROM {table} LIMIT {int(limit)}", conn)
            if df_sample.empty:
                continue
            out_path = os.path.join(debug_dir, f"{table}_sample_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv")
            df_sample.to_csv(out_path, index=False, encoding=models.CSV_ENCODING)
        except Exception:
            continue


def stream_read_and_write_csv(
    file: str,
    sheet: str,
    target_columns: List[str],
    temp_csv_path: str,
    fname: str,
    sheet_name: str,
    cast_stats_local: List[Dict[str, Any]],
    cast_failures_local: List[Dict[str, Any]],
    process_time: str,
    tax_text_to_zero: bool,
    stream_chunk_size: int,
    errors_list: Optional[List[Dict[str, Any]]] = None,
) -> int:
    try:
        import psutil

        available_mem_mb = psutil.virtual_memory().available / (1024 * 1024)
        dynamic_chunk_size = max(5000, min(100000, int(available_mem_mb * 0.1 * 1024)))
    except ImportError:
        dynamic_chunk_size = stream_chunk_size

    try:
        wb = load_workbook(filename=file, read_only=True, data_only=True)
    except Exception as e:
        if errors_list is not None:
            errors_list.append({"file": fname, "sheet": sheet_name, "stage": "open_workbook", "error_type": type(e).__name__, "message": str(e)})
        return 0

    if sheet not in wb.sheetnames:
        return 0

    ws = wb[sheet]
    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=True))]
    batch = []
    written = 0
    filtered_empty_rows = 0
    cols = header_row

    for _, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if all(cell is None or (isinstance(cell, str) and not cell.strip()) for cell in row):
            filtered_empty_rows += 1
            continue
        row_dict = {cols[j]: row[j] if j < len(row) else None for j in range(len(cols))}
        batch.append(row_dict)

        if len(batch) >= dynamic_chunk_size:
            df_chunk = pd.DataFrame(batch)
            before_filter = len(df_chunk)
            df_chunk = df_chunk.replace("", np.nan).dropna(how="all")
            filtered_empty_rows += before_filter - len(df_chunk)
            if len(df_chunk) == 0:
                batch = []
                continue
            try:
                df_chunk = cast_and_record(df_chunk, fname, sheet_name, cast_stats_local, cast_failures_local, tax_text_to_zero)
            except Exception as e:
                if errors_list is not None:
                    errors_list.append({"file": fname, "sheet": sheet_name, "stage": "stream_cast_chunk", "chunk_rows": len(batch), "error_type": type(e).__name__, "message": str(e)})
                batch = []
                continue
            df_chunk[models.AUDIT_SRC_FILE_COL] = fname
            df_chunk[models.AUDIT_IMPORT_TIME_COL] = process_time
            df_chunk = df_chunk.reindex(columns=list(target_columns))
            try:
                df_chunk.to_csv(temp_csv_path, mode="a", header=not os.path.exists(temp_csv_path), index=False, encoding=models.CSV_ENCODING)
            except Exception as e:
                if errors_list is not None:
                    errors_list.append({"file": fname, "sheet": sheet_name, "stage": "write_csv", "error_type": type(e).__name__, "message": str(e)})
                batch = []
                continue
            written += len(df_chunk)
            batch = []

    if batch:
        df_chunk = pd.DataFrame(batch)
        before_filter = len(df_chunk)
        df_chunk = df_chunk.replace("", np.nan).dropna(how="all")
        filtered_empty_rows += before_filter - len(df_chunk)
        if len(df_chunk) > 0:
            try:
                df_chunk = cast_and_record(df_chunk, fname, sheet_name, cast_stats_local, cast_failures_local, tax_text_to_zero)
            except Exception as e:
                if errors_list is not None:
                    errors_list.append({"file": fname, "sheet": sheet_name, "stage": "stream_cast_chunk", "chunk_rows": len(batch), "error_type": type(e).__name__, "message": str(e)})
            else:
                df_chunk[models.AUDIT_SRC_FILE_COL] = fname
                df_chunk[models.AUDIT_IMPORT_TIME_COL] = process_time
                df_chunk = df_chunk.reindex(columns=list(target_columns))
                try:
                    df_chunk.to_csv(temp_csv_path, mode="a", header=not os.path.exists(temp_csv_path), index=False, encoding=models.CSV_ENCODING)
                except Exception as e:
                    if errors_list is not None:
                        errors_list.append({"file": fname, "sheet": sheet_name, "stage": "write_csv_final", "error_type": type(e).__name__, "message": str(e)})
                else:
                    written += len(df_chunk)

    try:
        wb.close()
    except Exception:
        pass

    if filtered_empty_rows > 0:
        pass

    return written


def process_single_sheet(
    excel_file: str,
    sheet_name: str,
    handler: SheetTypeMapping,
    temp_dir: str,
    file_name: str,
    process_time: str,
    cast_stats_local: List[Dict[str, Any]],
    cast_failures_local: List[Dict[str, Any]],
    errors_list: List[Dict[str, Any]],
    df_queue: Optional[Any],
    use_csv_fallback: bool,
    use_streaming: bool,
    tax_text_to_zero: bool,
    stream_chunk_size: int,
) -> Tuple[int, str, str]:
    if handler is None:
        return 0, "ignored", ""

    rows_written = 0
    temp_csv_path = os.path.join(temp_dir, f"{handler.table_prefix}{file_name}__{sheet_name}__{uuid.uuid4().hex}.csv")
    classification = handler.classification

    try:
        if use_streaming:
            rows_written = stream_read_and_write_csv(
                excel_file,
                sheet_name,
                handler.target_columns,
                temp_csv_path,
                file_name,
                sheet_name,
                cast_stats_local,
                cast_failures_local,
                process_time,
                tax_text_to_zero,
                stream_chunk_size,
                errors_list=errors_list,
            )
            return rows_written, classification, temp_csv_path

        df = read_excel_with_engine(excel_file, sheet_name=sheet_name)
        df, rows_written = normalize_sheet_dataframe(
            df,
            sheet_name,
            file_name,
            process_time,
            handler.target_columns,
            lambda df_, fname, sheet, cs, cf: cast_and_record(df_, fname, sheet, cs, cf, tax_text_to_zero),
            cast_stats_local,
            cast_failures_local,
            errors_list,
            extract_year=True,
        )

        queued, msg = write_to_csv_or_queue(df, handler.target_table, temp_csv_path, df_queue, use_csv_fallback)
        if queued:
            return rows_written, classification, "queued"
        return rows_written, classification, temp_csv_path

    except MemoryError:
        rows_written = stream_read_and_write_csv(
            excel_file,
            sheet_name,
            handler.target_columns,
            temp_csv_path,
            file_name,
            sheet_name,
            cast_stats_local,
            cast_failures_local,
            process_time,
            tax_text_to_zero,
            stream_chunk_size,
            errors_list=errors_list,
        )
        return rows_written, classification, temp_csv_path
    except Exception as e:
        errors_list.append({"file": file_name, "sheet": sheet_name, "stage": "read", "error_type": type(e).__name__, "message": str(e)})
        return 0, "error", ""


def process_file_worker(args) -> Dict[str, Any]:
    (
        runtime,
        config,
        tax_text_to_zero,
        file,
        meta,
        temp_dir_root,
        process_time,
        detail_columns,
        header_columns,
        summary_columns,
        special_columns,
        stream_chunk_size,
    ) = args

    fname = os.path.basename(file)
    result = {"temp_csvs": [], "cast_stats_path": None, "cast_failures_path": None, "sheet_manifest": []}
    cast_stats_local: List[Dict[str, Any]] = []
    cast_failures_local: List[Dict[str, Any]] = []
    local_errors: List[Dict[str, Any]] = []
    temp_dir = ensure_worker_temp_dir(temp_dir_root)

    use_streaming_for_this_file = False
    if config and hasattr(config, "get"):
        try:
            enabled = config.get("performance", "memory_monitoring", "enabled", default=True)
            if bool(enabled):
                use_streaming_for_this_file = should_use_streaming_for_file(file, config)
        except Exception:
            use_streaming_for_this_file = False

    if str(file).lower().endswith(".xls"):
        use_streaming_for_this_file = False

    try:
        engine = "xlrd" if str(file).lower().endswith(".xls") else None
        with pd.ExcelFile(file, engine=engine) as xl:
            for sheet in xl.sheet_names:
                classification = "ignored"
                target_table = ""
                rows_written = 0
                try:
                    if sheet in meta.get("special_sheets", {}):
                        suffix = meta["special_sheets"][sheet]
                        target_cols = special_columns.get(suffix, [])
                        target_table = f"ODS_{runtime.business_tag}_SPECIAL_{suffix}"
                        temp_csv = os.path.join(temp_dir, f"{suffix}__{fname}__{sheet}__{uuid.uuid4().hex}.csv")
                        try:
                            if use_streaming_for_this_file:
                                rows_written = stream_read_and_write_csv(
                                    file,
                                    sheet,
                                    target_cols,
                                    temp_csv,
                                    fname,
                                    sheet,
                                    cast_stats_local,
                                    cast_failures_local,
                                    process_time,
                                    tax_text_to_zero,
                                    stream_chunk_size,
                                    errors_list=local_errors,
                                )
                            else:
                                df = read_excel_with_engine(file, sheet_name=sheet)
                                df = cast_and_record(df, fname, sheet, cast_stats_local, cast_failures_local, tax_text_to_zero)
                                df = add_audit_columns(df, fname, process_time)
                                if models.INVOICE_DATE_COL in df.columns:
                                    df = add_invoice_year_column(df)
                                df = filter_dataframe_columns(df, list(target_cols))
                                save_dataframe_to_csv(df, temp_csv)
                                rows_written = len(df)
                        except MemoryError:
                            rows_written = stream_read_and_write_csv(
                                file,
                                sheet,
                                target_cols,
                                temp_csv,
                                fname,
                                sheet,
                                cast_stats_local,
                                cast_failures_local,
                                process_time,
                                tax_text_to_zero,
                                stream_chunk_size,
                                errors_list=local_errors,
                            )
                        result["temp_csvs"].append({"path": temp_csv, "target_table": target_table, "rows": rows_written})
                        classification = f"special_{suffix.lower()}"
                    elif sheet in meta.get("summary_sheets", []):
                        target_cols = summary_columns
                        target_table = f"ODS_{runtime.business_tag}_DETAIL"
                        temp_csv = os.path.join(temp_dir, f"DETAIL__{fname}__{sheet}__{uuid.uuid4().hex}.csv")
                        try:
                            if use_streaming_for_this_file:
                                rows_written = stream_read_and_write_csv(
                                    file,
                                    sheet,
                                    target_cols,
                                    temp_csv,
                                    fname,
                                    sheet,
                                    cast_stats_local,
                                    cast_failures_local,
                                    process_time,
                                    tax_text_to_zero,
                                    stream_chunk_size,
                                    errors_list=local_errors,
                                )
                            else:
                                df = read_excel_with_engine(file, sheet_name=sheet)
                                df = cast_and_record(df, fname, sheet, cast_stats_local, cast_failures_local, tax_text_to_zero)
                                df = add_audit_columns(df, fname, process_time)
                                if models.INVOICE_DATE_COL in df.columns:
                                    df = add_invoice_year_column(df)
                                df = filter_dataframe_columns(df, list(target_cols))
                                save_dataframe_to_csv(df, temp_csv)
                                rows_written = len(df)
                        except MemoryError:
                            rows_written = stream_read_and_write_csv(
                                file,
                                sheet,
                                target_cols,
                                temp_csv,
                                fname,
                                sheet,
                                cast_stats_local,
                                cast_failures_local,
                                process_time,
                                tax_text_to_zero,
                                stream_chunk_size,
                                errors_list=local_errors,
                            )
                        result["temp_csvs"].append({"path": temp_csv, "target_table": target_table, "rows": rows_written})
                        try:
                            key_cols = [c for c in models.INVOICE_KEY_COLS if c in df.columns]
                            if key_cols:
                                result["summary_keys"] = df[key_cols].drop_duplicates().to_dict(orient="records")
                        except Exception:
                            pass
                        classification = "summary"
                    elif sheet in meta.get("detail_sheets", []):
                        target_cols = detail_columns
                        target_table = f"ODS_{runtime.business_tag}_TEMP_TRANSIT"
                        temp_csv = os.path.join(temp_dir, f"TEMP_TRANSIT__{fname}__{sheet}__{uuid.uuid4().hex}.csv")
                        try:
                            if use_streaming_for_this_file:
                                rows_written = stream_read_and_write_csv(
                                    file,
                                    sheet,
                                    target_cols,
                                    temp_csv,
                                    fname,
                                    sheet,
                                    cast_stats_local,
                                    cast_failures_local,
                                    process_time,
                                    tax_text_to_zero,
                                    stream_chunk_size,
                                    errors_list=local_errors,
                                )
                            else:
                                df = read_excel_with_engine(file, sheet_name=sheet)
                                df = cast_and_record(df, fname, sheet, cast_stats_local, cast_failures_local, tax_text_to_zero)
                                df = add_audit_columns(df, fname, process_time)
                                if models.INVOICE_DATE_COL in df.columns:
                                    df = add_invoice_year_column(df)
                                df = filter_dataframe_columns(df, list(target_cols))
                                save_dataframe_to_csv(df, temp_csv)
                                rows_written = len(df)
                        except MemoryError:
                            rows_written = stream_read_and_write_csv(
                                file,
                                sheet,
                                target_cols,
                                temp_csv,
                                fname,
                                sheet,
                                cast_stats_local,
                                cast_failures_local,
                                process_time,
                                tax_text_to_zero,
                                stream_chunk_size,
                                errors_list=local_errors,
                            )
                        result["temp_csvs"].append({"path": temp_csv, "target_table": target_table, "rows": rows_written})
                        classification = "detail"
                    elif sheet in meta.get("header_sheets", []):
                        target_cols = header_columns
                        target_table = f"ODS_{runtime.business_tag}_HEADER"
                        temp_csv = os.path.join(temp_dir, f"HEADER__{fname}__{sheet}__{uuid.uuid4().hex}.csv")
                        try:
                            if use_streaming_for_this_file:
                                rows_written = stream_read_and_write_csv(
                                    file,
                                    sheet,
                                    target_cols,
                                    temp_csv,
                                    fname,
                                    sheet,
                                    cast_stats_local,
                                    cast_failures_local,
                                    process_time,
                                    tax_text_to_zero,
                                    stream_chunk_size,
                                    errors_list=local_errors,
                                )
                            else:
                                df = read_excel_with_engine(file, sheet_name=sheet)
                                df = cast_and_record(df, fname, sheet, cast_stats_local, cast_failures_local, tax_text_to_zero)
                                df = add_audit_columns(df, fname, process_time)
                                if models.INVOICE_DATE_COL in df.columns:
                                    df = add_invoice_year_column(df)
                                df = filter_dataframe_columns(df, list(target_cols))
                                save_dataframe_to_csv(df, temp_csv)
                                rows_written = len(df)
                        except MemoryError:
                            rows_written = stream_read_and_write_csv(
                                file,
                                sheet,
                                target_cols,
                                temp_csv,
                                fname,
                                sheet,
                                cast_stats_local,
                                cast_failures_local,
                                process_time,
                                tax_text_to_zero,
                                stream_chunk_size,
                                errors_list=local_errors,
                            )
                        result["temp_csvs"].append({"path": temp_csv, "target_table": target_table, "rows": rows_written})
                        classification = "header"
                except Exception:
                    classification = "error"
                cols = meta.get("sheet_info", {}).get(sheet, [])
                result["sheet_manifest"].append({"file": fname, "sheet": sheet, "classification": classification, "columns": ";".join(cols), "target_table": target_table, "rows": rows_written})

        if cast_stats_local:
            cs_path = os.path.join(temp_dir, f"cast_stats_{uuid.uuid4().hex}.csv")
            pd.DataFrame(cast_stats_local).to_csv(cs_path, index=False, encoding=models.CSV_ENCODING)
            result["cast_stats_path"] = cs_path
        if cast_failures_local:
            cf_path = os.path.join(temp_dir, f"cast_failures_{uuid.uuid4().hex}.csv")
            pd.concat(cast_failures_local, ignore_index=True).to_csv(cf_path, index=False, encoding=models.CSV_ENCODING)
            result["cast_failures_path"] = cf_path
    except Exception:
        pass

    if local_errors:
        result["errors"] = local_errors
    return result


def merge_temp_csvs_to_db(
    temp_dir: str,
    conn: sqlite3.Connection,
    table_columns_map: Dict[str, List[str]],
    csv_chunk_size: int,
    business_tag: str,
    error_logs: Optional[List[Dict[str, Any]]] = None,
) -> None:
    cursor = conn.cursor()
    try:
        cursor.execute("PRAGMA journal_mode=WAL")
        cursor.execute("PRAGMA synchronous=OFF")
    except Exception:
        pass

    temp_files = []
    for root, _, files in os.walk(temp_dir):
        for f in files:
            if f.lower().endswith(".csv"):
                temp_files.append(os.path.join(root, f))
    grouped: Dict[str, List[str]] = {}
    for f in temp_files:
        bn = os.path.basename(f)
        if bn.startswith("cast_stats_") or bn.startswith("cast_failures_"):
            continue
        prefix = bn.split(models.FILE_SPLIT_DELIMITER, 1)[0] if models.FILE_SPLIT_DELIMITER in bn else None
        assigned = None
        if prefix:
            if prefix in ("TEMP_TRANSIT", "TEMP"):
                assigned = f"ODS_{business_tag}_TEMP_TRANSIT"
            elif prefix == "HEADER":
                assigned = f"ODS_{business_tag}_HEADER"
            elif prefix == "DETAIL":
                assigned = f"ODS_{business_tag}_DETAIL"
            else:
                for tbl in table_columns_map.keys():
                    if tbl.upper().endswith("_" + prefix.upper()) or tbl.upper().endswith(prefix.upper()):
                        assigned = tbl
                        break
        if not assigned:
            try:
                df_sample = read_csv_with_encoding_detection(f, nrows=0)
                cols = set(df_sample.columns.tolist())
                for tbl, tbl_cols in table_columns_map.items():
                    if len(cols & set(tbl_cols)) > 0:
                        assigned = tbl
                        break
            except Exception:
                assigned = None
        if not assigned:
            assigned = f"ODS_{business_tag}_TEMP_TRANSIT"
        grouped.setdefault(assigned, []).append(f)

    for tbl, files in grouped.items():
        try:
            cursor.execute("BEGIN IMMEDIATE")
            for f in files:
                try:
                    for chunk_no, chunk in enumerate(read_csv_with_encoding_detection(f, chunksize=csv_chunk_size)):
                        try:
                            chunk.to_sql(tbl, conn, if_exists="append", index=False, method="multi", chunksize=500)
                        except Exception as ce:
                            if error_logs is not None:
                                error_logs.append({"stage": "merge_chunk", "file": f, "target_table": tbl, "chunk_no": chunk_no, "error_type": type(ce).__name__, "message": str(ce)})
                except Exception as fe:
                    if error_logs is not None:
                        error_logs.append({"stage": "read_temp_csv", "file": f, "target_table": tbl, "error_type": type(fe).__name__, "message": str(fe)})
            conn.commit()
        except Exception as e:
            try:
                conn.rollback()
            except Exception:
                pass
            if error_logs is not None:
                error_logs.append({"stage": "merge_group", "target_table": tbl, "files": files, "error_type": type(e).__name__, "message": str(e)})

    try:
        cursor.execute("PRAGMA synchronous=NORMAL")
    except Exception:
        pass


def _reorder_header_columns(columns: List[str], business_tag: str) -> List[str]:
    """
    按照指定的标准顺序重新排列 ODS_VAT_INV_HEADER 表的字段。
    
    标准字段顺序（按此顺序排列）：
    1. 发票代码
    2. 发票号码
    3. 数电发票号码
    4. 销方识别号
    5. 销方名称
    6. 购方识别号
    7. 购买方名称
    8. 开票日期
    9. 金额
    10. 税额
    11. 价税合计
    12. 发票来源
    13. 发票票种
    14. 发票状态
    15. 是否正数发票
    16. 发票风险等级
    17. 开票人
    18. 备注
    
    其他字段会排在后面（保持原有相对顺序）。
    """
    # 定义标准字段顺序（大小写敏感）
    # 标准字段顺序 + 必须补齐的技术字段
    standard_order = [
        "header_uuid", "source_system", "created_at",  # 技术字段，优先补齐
        "发票代码",
        "发票号码",
        "数电发票号码",
        "销方识别号",
        "销方名称",
        "购方识别号",
        "购买方名称",
        "开票日期",
        "金额",
        "税额",
        "价税合计",
        "发票来源",
        "发票票种",
        "发票状态",
        "是否正数发票",
        "发票风险等级",
        "开票人",
        "备注",
    ]
    
    input_set = set(columns)
    ordered_columns = []

    # 补齐缺失的技术字段（如header_uuid、source_system、created_at）
    for field in ["header_uuid", "source_system", "created_at"]:
        if field not in input_set:
            ordered_columns.append(field)
    # 按标准顺序添加已有字段
    for field in standard_order:
        if field in input_set:
            ordered_columns.append(field)
            input_set.remove(field)
    # 其他字段保持原有顺序
    remaining = [col for col in columns if col in input_set]
    ordered_columns.extend(remaining)
    return ordered_columns


def _prepare_ods_tables(
    conn: sqlite3.Connection,
    detail_columns: List[str],
    header_columns: List[str],
    summary_columns: List[str],
    special_columns: Dict[str, List[str]],
    business_tag: str,
) -> None:
    cursor = conn.cursor()
    # 只允许业务标签为年份（如2021、2022、2023等）才创建明细/表头表，彻底移除VAT_INV标签的表
    if business_tag.isdigit():
        # 用户指定ODS表头字段顺序和字段集，严格对齐
        strict_header_columns = [
            "header_uuid","source_system","created_at","created_by","updated_at","updated_by","import_batch_id","sync_status","clean_status","detail_total_amount","is_balanced","balance_diff","balance_tolerance","balance_check_time","balance_check_by","balance_notes","related_blue_invoice_uuid","fpdm","fphm","sdfphm","xfsbh","xfmc","gfsbh","gfmc","kprq","invoice_date","invoice_time","je","se","jshj","fply","fppz","fpzt","sfzsfp","fpfxdj","kpr","bz"
        ]
        strict_detail_columns = [
            "detail_uuid","header_uuid","logic_line_no","updated_at","updated_by","import_batch_id","source_system","sync_status","clean_status","fpdm","fphm","sdfphm","invoice_date","hwlwmc","ggxh","dw","sl","dj","je","slv","se","jshj"
        ]
        # 先强制删除旧表，彻底覆盖
        cursor.execute(f"DROP TABLE IF EXISTS ODS_VAT_INV_DETAIL_FULL_{business_tag}")
        cursor.execute(f"DROP TABLE IF EXISTS ODS_VAT_INV_HEADER_FULL_{business_tag}")
        pd.DataFrame(columns=strict_detail_columns).to_sql(f"ODS_VAT_INV_DETAIL_FULL_{business_tag}", conn, if_exists="replace", index=False, method="multi")
        pd.DataFrame(columns=list(summary_columns)).to_sql(f"ODS_VAT_INV_DETAIL_FULL_{business_tag}", conn, if_exists="replace", index=False, method="multi")
        pd.DataFrame(columns=strict_header_columns).to_sql(f"ODS_VAT_INV_HEADER_FULL_{business_tag}", conn, if_exists="replace", index=False, method="multi")

    # 预创建特殊表以保持模式与扫描列的同步（避免追加时的缺失列错误）
    for suffix, cols in special_columns.items():
        table_name = f"ODS_VAT_INV_SPECIAL_{business_tag}_{suffix}"
        cursor.execute(f"DROP TABLE IF EXISTS {table_name}")
        pd.DataFrame(columns=list(cols)).to_sql(table_name, conn, if_exists="replace", index=False, method="multi")


def _export_ods_manifest(runtime: RuntimeContext, sheet_manifest, cast_stats, cast_failures, process_time: str, output_dir: str, logger) -> None:
    if sheet_manifest:
        manifest_path = os.path.join(output_dir, generate_manifest_filename(models.MANIFEST_PREFIX, process_time))
        save_dataframe_to_csv(pd.DataFrame(sheet_manifest), manifest_path)
        logger.info(f"Sheet清单已导出: {manifest_path}")

    if cast_stats:
        df_cast = pd.DataFrame(cast_stats)
        cast_manifest_path = os.path.join(output_dir, generate_manifest_filename(models.CAST_STATS_PREFIX, process_time))
        save_dataframe_to_csv(df_cast, cast_manifest_path)
        logger.info(f"导出类型转换清单: {cast_manifest_path}")
    else:
        logger.info("未发现类型转换操作或无需记录。")

    if cast_failures:
        try:
            df_fail = pd.concat(cast_failures, ignore_index=True)
            df_fail_limited = df_fail.groupby("column").head(runtime.max_failure_sample_per_col).reset_index(drop=True)
            fail_manifest_path = os.path.join(output_dir, generate_manifest_filename(models.CAST_FAILURES_PREFIX, process_time))
            save_dataframe_to_csv(df_fail_limited, fail_manifest_path)
            _progress(f"导出类型转换失败样本: {fail_manifest_path}")
        except Exception as e:
            logger.warning(f"导出类型转换失败样本失败: {e}")


def _import_ods_data(
    runtime: RuntimeContext,
    conn: sqlite3.Connection,
    excel_files: List[str],
    files_meta: Dict[str, Any],
    detail_columns: List[str],
    header_columns: List[str],
    summary_columns: List[str],
    special_columns: Dict[str, List[str]],
    process_time: str,
    config: Optional[Any],
    logger,
) -> Dict[str, Any]:
    sheet_manifest: List[Dict[str, Any]] = []
    processed_files = set()
    read_failed_files: List[str] = []
    cast_stats: List[Any] = []
    cast_failures: List[pd.DataFrame] = []
    error_logs: List[Dict[str, Any]] = []
    temp_root = None

    cursor = conn.cursor()
    total_sheets = sum(len(meta.get("sheet_info", {})) for meta in files_meta.values()) if files_meta else 0
    processed_sheets = 0
    _progress(f"准备逐文件逐 sheet 处理：共 {len(excel_files)} 个文件，{total_sheets} 个 sheet。")

    if runtime.enable_parallel_import:
        io_cfg = {}
        if config and hasattr(config, "get"):
            io_cfg = config.get("performance", "io_throttle", default={}) or {}
        io_enabled = io_cfg.get("enabled", True)
        io_threshold = io_cfg.get("busy_threshold_percent", 75)
        io_sample = io_cfg.get("sample_seconds", 0.25)
        io_reduce = io_cfg.get("reduce_factor", 0.5)
        io_min_workers = io_cfg.get("min_workers", 1)

        disk_busy = measure_disk_busy_percent(sample_seconds=io_sample) if io_enabled else None
        if disk_busy is not None:
            _debug_var("disk_busy_percent", f"{disk_busy:.1f}%")

        dynamic_worker_count = calculate_optimal_workers(
            excel_files,
            runtime.worker_count,
            disk_busy_percent=disk_busy,
            io_threshold=io_threshold,
            reduce_factor=io_reduce,
            min_workers=io_min_workers,
        )
        _progress(f"启用并行导入：使用 {dynamic_worker_count} 个 worker 处理文件（使用 CSV 临时文件方案）")

        temp_root = os.path.join(runtime.output_dir, models.TEMP_FILE_PREFIX, format_timestamp_for_filename(process_time))
        if os.path.exists(temp_root):
            shutil.rmtree(temp_root)
        os.makedirs(temp_root, exist_ok=True)

        cleanup_old_temp_files(os.path.join(runtime.output_dir, models.TEMP_FILE_PREFIX))

        register_cleanup(temp_root)

        worker_args = []
        for file in excel_files:
            fname = os.path.basename(file)
            if fname.startswith("~$"):
                read_failed_files.append(fname)
                continue
            meta = files_meta.get(fname)
            if not meta:
                read_failed_files.append(fname)
                error_logs.append({"file": fname, "stage": "metadata", "error_type": "MetadataMissing", "message": "元数据扫描失败，跳过该文件"})
                continue
            meta = meta or {"sheet_info": {}, "detail_sheets": [], "header_sheets": [], "summary_sheets": [], "special_sheets": {}}
            worker_args.append(
                (
                    runtime,
                    config,
                    runtime.tax_text_to_zero,
                    file,
                    meta,
                    temp_root,
                    process_time,
                    detail_columns,
                    header_columns,
                    summary_columns,
                    special_columns,
                    runtime.stream_chunk_size,
                )
            )

        results = []
        worker_timer = PerformanceTimer("Worker并行执行")
        worker_timer.__enter__()
        with multiprocessing.Pool(dynamic_worker_count) as pool:
            for res in pool.imap_unordered(process_file_worker, worker_args):
                results.append(res)
                for entry in res.get("sheet_manifest", []):
                    sheet_manifest.append(entry)
                    processed_sheets += 1
                    _progress(
                        f"[{processed_sheets}/{total_sheets}] {entry['file']} - {entry['sheet']}: {entry['classification']} -> {entry.get('target_table') or '-'} ({entry.get('rows') or '-'} rows)"
                    )
                if any(e["classification"] != "ignored" and e["classification"] != "error" for e in res.get("sheet_manifest", [])):
                    processed_files.add(os.path.basename(res.get("sheet_manifest", [{}])[0].get("file", "")))
                if res.get("cast_stats_path"):
                    try:
                        df_cs = pd.read_csv(res["cast_stats_path"], encoding=models.CSV_ENCODING)
                        cast_stats.append(df_cs.to_dict(orient="records"))
                    except Exception:
                        pass
                if res.get("cast_failures_path"):
                    try:
                        df_cf = pd.read_csv(res["cast_failures_path"], encoding=models.CSV_ENCODING)
                        cast_failures.append(df_cf)
                    except Exception:
                        pass
                if res.get("errors"):
                    error_logs.extend(res.get("errors"))

        worker_timer.__exit__()
        worker_timer.log()

        # 对 header_columns 进行排序，应用标准字段顺序
        sorted_header_columns = _reorder_header_columns(header_columns, runtime.business_tag)
        
        table_columns_map = {
            f"ODS_{runtime.business_tag}_TEMP_TRANSIT": detail_columns,
            f"ODS_{runtime.business_tag}_HEADER": sorted_header_columns,
            f"ODS_{runtime.business_tag}_DETAIL": summary_columns,
        }
        for suffix, cols in special_columns.items():
            table_columns_map[f"ODS_{runtime.business_tag}_SPECIAL_{suffix}"] = cols
        merge_conn = sqlite3.connect(runtime.db_path)
        merge_temp_csvs_to_db(temp_root, merge_conn, table_columns_map, runtime.csv_chunk_size, runtime.business_tag, error_logs)
        merge_conn.commit()
        merge_conn.close()
        conn = sqlite3.connect(runtime.db_path)

        if runtime.debug_mode:
            try:
                from vat_audit_pipeline.core.processors.ods_processor import export_debug_samples  # avoid circular
            except Exception:
                pass

        flat_cs: List[Dict[str, Any]] = []
        for item in cast_stats:
            if isinstance(item, list):
                flat_cs.extend(item)
            elif isinstance(item, pd.DataFrame):
                flat_cs.extend(item.to_dict(orient="records"))
            else:
                try:
                    flat_cs.extend(list(item))
                except Exception:
                    pass
        cast_stats = flat_cs
        if cast_failures:
            cast_failures = [df for df in cast_failures if isinstance(df, pd.DataFrame)]
    else:
        for file in excel_files:
            fname = os.path.basename(file)
            if fname.startswith("~$"):
                read_failed_files.append(fname)
                continue
            meta = files_meta.get(fname)
            if not meta:
                read_failed_files.append(fname)
                error_logs.append({"file": fname, "stage": "metadata", "error_type": "MetadataMissing", "message": "元数据扫描失败，跳过该文件"})
                continue
            file_success = False
            try:
                cursor.execute("BEGIN IMMEDIATE")
                engine = "xlrd" if str(file).lower().endswith(".xls") else None
                xl = pd.ExcelFile(file, engine=engine)
                meta = meta or {"sheet_info": {}, "detail_sheets": [], "header_sheets": [], "summary_sheets": []}
                for sheet in xl.sheet_names:
                    cols = meta["sheet_info"].get(sheet, []) if meta else []
                    classification = "ignored"
                    target_table = ""
                    rows = None
                    try:
                        if sheet in meta.get("special_sheets", {}):
                            suffix = meta["special_sheets"][sheet]
                            df = read_excel_with_engine(file, sheet_name=sheet)
                            df = cast_and_record(df, fname, sheet, cast_stats, cast_failures, runtime.tax_text_to_zero)
                            df[models.AUDIT_SRC_FILE_COL] = fname
                            df[models.AUDIT_IMPORT_TIME_COL] = process_time
                            if models.INVOICE_DATE_COL in df.columns:
                                df[models.INVOICE_YEAR_COL] = df[models.INVOICE_DATE_COL].astype(str).str[:4]
                            else:
                                df[models.INVOICE_YEAR_COL] = None
                            df = df.reindex(columns=list(special_columns.get(suffix, [])))
                            target_table = f"ODS_VAT_INV_SPECIAL_{runtime.business_tag}_{suffix}"
                            df.to_sql(target_table, conn, if_exists="append", index=False, method="multi", chunksize=500)
                            rows = len(df)
                            classification = f"special_{suffix.lower()}"
                            file_success = True
                            del df
                        elif sheet in meta.get("summary_sheets", []):
                            df = read_excel_with_engine(file, sheet_name=sheet)
                            df = cast_and_record(df, fname, sheet, cast_stats, cast_failures, runtime.tax_text_to_zero)
                            df[models.AUDIT_SRC_FILE_COL] = fname
                            df[models.AUDIT_IMPORT_TIME_COL] = process_time
                            if models.INVOICE_DATE_COL in df.columns:
                                df[models.INVOICE_YEAR_COL] = df[models.INVOICE_DATE_COL].astype(str).str[:4]
                            else:
                                df[models.INVOICE_YEAR_COL] = None
                            df = df.reindex(columns=list(summary_columns))
                            target_table = f"ODS_VAT_INV_DETAIL_FULL_{runtime.business_tag}"
                            df.to_sql(target_table, conn, if_exists="append", index=False, method="multi", chunksize=500)
                            rows = len(df)
                            key_cols = [c for c in models.INVOICE_KEY_COLS if c in df.columns]
                            if key_cols:
                                files_meta[fname]["summary_df"] = df[key_cols].drop_duplicates()
                                files_meta[fname]["summary_key_cols"] = key_cols
                            classification = "summary"
                            file_success = True
                            del df
                        elif sheet in meta.get("detail_sheets", []):
                            df = read_excel_with_engine(file, sheet_name=sheet)
                            df = cast_and_record(df, fname, sheet, cast_stats, cast_failures, runtime.tax_text_to_zero)
                            df[models.AUDIT_SRC_FILE_COL] = fname
                            df[models.AUDIT_IMPORT_TIME_COL] = process_time
                            if models.INVOICE_DATE_COL in df.columns:
                                df[models.INVOICE_YEAR_COL] = df[models.INVOICE_DATE_COL].astype(str).str[:4]
                            else:
                                df[models.INVOICE_YEAR_COL] = None
                            strict_detail_columns = [
                                "detail_uuid","header_uuid","logic_line_no","updated_at","updated_by","import_batch_id","source_system","sync_status","clean_status","fpdm","fphm","sdfphm","invoice_date","hwlwmc","ggxh","dw","sl","dj","je","slv","se","jshj"
                            ]
                            df = df.reindex(columns=strict_detail_columns)
                            target_table = f"ODS_VAT_INV_DETAIL_FULL_{runtime.business_tag}"
                            df.to_sql(target_table, conn, if_exists="append", index=False, method="multi", chunksize=500)
                            rows = len(df)
                            del df
                            classification = "detail"
                            file_success = True
                        elif sheet in meta.get("header_sheets", []):
                            df = read_excel_with_engine(file, sheet_name=sheet)
                            df = cast_and_record(df, fname, sheet, cast_stats, cast_failures, runtime.tax_text_to_zero)
                            df[models.AUDIT_SRC_FILE_COL] = fname
                            df[models.AUDIT_IMPORT_TIME_COL] = process_time
                            if models.INVOICE_DATE_COL in df.columns:
                                df[models.INVOICE_YEAR_COL] = df[models.INVOICE_DATE_COL].astype(str).str[:4]
                            else:
                                df[models.INVOICE_YEAR_COL] = None
                            strict_header_columns = [
                                "header_uuid","source_system","created_at","created_by","updated_at","updated_by","import_batch_id","sync_status","clean_status","detail_total_amount","is_balanced","balance_diff","balance_tolerance","balance_check_time","balance_check_by","balance_notes","related_blue_invoice_uuid","fpdm","fphm","sdfphm","xfsbh","xfmc","gfsbh","gfmc","kprq","invoice_date","invoice_time","je","se","jshj","fply","fppz","fpzt","sfzsfp","fpfxdj","kpr","bz"
                            ]
                            df = df.reindex(columns=strict_header_columns)
                            target_table = f"ODS_VAT_INV_HEADER_FULL_{runtime.business_tag}"
                            df.to_sql(target_table, conn, if_exists="append", index=False, method="multi", chunksize=500)
                            rows = len(df)
                            del df
                            classification = "header"
                            file_success = True
                        else:
                            classification = "ignored"
                    except Exception as e:
                        try:
                            conn.rollback()
                        except Exception:
                            pass
                        err_entry = {"file": fname, "sheet": sheet, "stage": "read_or_write", "error_type": type(e).__name__, "message": str(e)}
                        error_logs.append(err_entry)
                        classification = "error"
                    sheet_manifest.append({"file": fname, "sheet": sheet, "classification": classification, "columns": ";".join(cols), "target_table": target_table, "rows": rows})
                    processed_sheets += 1
                    _progress(f"[{processed_sheets}/{total_sheets}] {fname} - {sheet}: {classification} -> {target_table or '-'} ({rows if rows is not None else '-'} rows)")
                if file_success:
                    try:
                        conn.commit()
                    except Exception:
                        pass
                    processed_files.add(fname)
                else:
                    try:
                        conn.rollback()
                    except Exception:
                        pass
                    read_failed_files.append(fname)
            except Exception as e:
                try:
                    conn.rollback()
                except Exception:
                    pass
                err_entry = {"file": fname, "stage": "file_read", "error_type": type(e).__name__, "message": str(e)}
                error_logs.append(err_entry)
                read_failed_files.append(fname)

    return {
        "sheet_manifest": sheet_manifest,
        "processed_files": processed_files,
        "read_failed_files": read_failed_files,
        "cast_stats": cast_stats,
        "cast_failures": cast_failures,
        "error_logs": error_logs,
        "temp_root": temp_root,
    }


def process_ods(
    runtime: RuntimeContext,
    logger,
    excel_files: List[str],
    files_meta: Dict[str, Any],
    detail_columns: List[str],
    header_columns: List[str],
    summary_columns: List[str],
    special_columns: Dict[str, List[str]],
    process_time: str,
    config: Optional[Any],
    conn: sqlite3.Connection,
) -> Dict[str, Any]:
    perf_timer = PerformanceTimer("ODS导入流程")
    perf_timer.__enter__()
    mem_monitor = MemoryMonitor()
    mem_monitor.start()

    _prepare_ods_tables(conn, detail_columns, header_columns, summary_columns, special_columns, runtime.business_tag)

    import_result = _import_ods_data(
        runtime,
        conn,
        excel_files,
        files_meta,
        detail_columns,
        header_columns,
        summary_columns,
        special_columns,
        process_time,
        config,
        logger,
    )

    total_files = len(excel_files)
    processed = len(import_result["processed_files"])
    scan_failed = len([f for f in files_meta.keys() if not files_meta.get(f)])
    read_failed = len(import_result["read_failed_files"])
    logger.info(f"导入汇总：共发现 {total_files} 个文件；成功导入 {processed} 个；列扫描失败 {scan_failed} 个；读取/写入失败 {read_failed} 个。")

    summary = {
        "total_files": total_files,
        "processed_files": processed,
        "scan_failed_files": ";".join([f for f in files_meta.keys() if not files_meta.get(f)]),
        "read_failed_files": ";".join(import_result["read_failed_files"]),
        "process_time": process_time,
    }

    _export_ods_manifest(runtime, import_result["sheet_manifest"], import_result["cast_stats"], import_result["cast_failures"], process_time, runtime.output_dir, logger)

    perf_timer.__exit__()
    perf_timer.log()
    mem_monitor.end()
    mem_monitor.log()

    return {
        "sheet_manifest": import_result["sheet_manifest"],
        "processed_files": import_result["processed_files"],
        "read_failed_files": import_result["read_failed_files"],
        "summary": summary,
        "cast_stats": import_result["cast_stats"],
        "cast_failures": import_result["cast_failures"],
        "error_logs": import_result["error_logs"],
        "temp_root": import_result.get("temp_root"),
    }